#!/usr/bin/env python3
"""Prepara planilha MR 2026 CONTATOS → CSV/XLSX para importação ZapMass."""
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    raise SystemExit("Instale openpyxl: pip install openpyxl") from e

DEFAULT_SRC = Path(r"C:\Users\xgame\Downloads\Cópia de MR 2026 _ CONTATOS.xlsx")
OUT_DIR = Path.home() / "Downloads"

COLUMNS = [
    "Nome",
    "Telefone",
    "Email",
    "Cidade",
    "Cargo (Igreja)",
    "Tags (separadas por ;)",
    "Observações",
]


def norm_header(h: str) -> str:
    return (
        str(h or "")
        .strip()
        .lower()
        .encode("ascii", "ignore")
        .decode("ascii")
    )


def parse_phone(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if re.search(r"[Ee][+\-]?\d+", s):
        try:
            s = str(int(round(float(s.replace(",", ".")))))
        except Exception:
            pass
    d = re.sub(r"\D", "", s)
    if not d:
        return ""
    if d.startswith("55") and len(d) >= 12:
        return d[:13]
    if len(d) in (10, 11):
        return "55" + d
    if len(d) >= 12:
        return ("55" + d.lstrip("0"))[:13]
    return d if len(d) >= 10 else ""


def clean_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return re.sub(r"\s+", " ", s)


def pick_name(a: str, b: str) -> str:
    a, b = clean_text(a), clean_text(b)
    if not a:
        return b
    if not b:
        return a
    return a if len(a) >= len(b) else b


def read_source(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    raw: list[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(x or "").strip() for x in rows[0]]
        hdr_lower = [norm_header(h) for h in hdr]

        name_idx = next(
            (i for i, h in enumerate(hdr_lower) if h in ("nome", "nico", "name")),
            0,
        )
        city_idx = next((i for i, h in enumerate(hdr_lower) if "cidade" in h), None)
        email_idx = next((i for i, h in enumerate(hdr_lower) if "mail" in h), None)
        role_idx = next((i for i, h in enumerate(hdr_lower) if "fun" in h), None)
        phone_idx = next(
            (i for i, h in enumerate(hdr_lower) if "telefone" in h or h == "phone"),
            None,
        )
        obs_idx = next((i for i, h in enumerate(hdr_lower) if "observ" in h), None)

        for row in rows[1:]:
            if not row or not any(row):
                continue
            vals = list(row)
            while len(vals) < len(hdr):
                vals.append(None)

            name = clean_text(vals[name_idx] if name_idx is not None else "")
            if name.lower() in ("nome", "nico", "name"):
                continue

            phone = parse_phone(vals[phone_idx] if phone_idx is not None else "")
            city = clean_text(vals[city_idx] if city_idx is not None else "")
            email = clean_text(vals[email_idx] if email_idx is not None else "")
            role = clean_text(vals[role_idx] if role_idx is not None else "")
            obs = clean_text(vals[obs_idx] if obs_idx is not None else "")

            if not name and not phone:
                continue

            raw.append(
                {
                    "sheet": sheet,
                    "nome": name or "Sem nome",
                    "telefone": phone,
                    "email": email,
                    "cidade": city,
                    "cargo": role,
                    "obs": obs,
                }
            )
    return raw


def merge_rows(raw: list[dict]) -> tuple[list[dict], dict]:
    by_phone: dict[str, dict] = {}
    no_phone: list[dict] = []

    for r in raw:
        phone = r["telefone"]
        if not phone or len(phone) < 12:
            no_phone.append(r)
            continue
        if phone not in by_phone:
            by_phone[phone] = {
                "Nome": r["nome"],
                "Telefone": phone,
                "Email": r["email"],
                "Cidade": r["cidade"],
                "Cargo (Igreja)": r["cargo"],
                "Tags (separadas por ;)": {"MR 2026", "Importado", r["sheet"]},
                "Observações": r["obs"],
            }
            continue
        cur = by_phone[phone]
        cur["Nome"] = pick_name(cur["Nome"], r["nome"])
        if not cur["Email"] and r["email"]:
            cur["Email"] = r["email"]
        if not cur["Cidade"] and r["cidade"]:
            cur["Cidade"] = r["cidade"]
        if not cur["Cargo (Igreja)"] and r["cargo"]:
            cur["Cargo (Igreja)"] = r["cargo"]
        cur["Tags (separadas por ;)"].add(r["sheet"])
        if r["obs"]:
            prev = cur["Observações"]
            cur["Observações"] = f"{prev}; {r['obs']}".strip("; ") if prev else r["obs"]

    out: list[dict] = []
    for cur in by_phone.values():
        tags = sorted(cur["Tags (separadas por ;)"], key=str.lower)
        out.append(
            {
                "Nome": cur["Nome"],
                "Telefone": cur["Telefone"],
                "Email": cur["Email"],
                "Cidade": cur["Cidade"],
                "Cargo (Igreja)": cur["Cargo (Igreja)"],
                "Tags (separadas por ;)": ";".join(tags),
                "Observações": cur["Observações"],
            }
        )
    out.sort(key=lambda x: (x["Nome"] or "").lower())

    stats = {
        "linhas_planilha": len(raw),
        "sem_telefone": len(no_phone),
        "exportados": len(out),
        "duplicatas_mescladas": len(raw) - len(no_phone) - len(out),
    }
    return out, stats


def write_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, delimiter=";")
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatos"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    wb.save(path)


def main() -> int:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.is_file():
        print(f"Arquivo não encontrado: {src}")
        return 1

    raw = read_source(src)
    rows, stats = merge_rows(raw)

    base = OUT_DIR / "MR_2026_CONTATOS_zapmass_import"
    csv_path = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")

    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)

    print(f"Origem: {src}")
    print(f"Abas processadas: HOMENS, MULHERES, JOVENS, SENIOR, COMUNIDADES")
    print(f"Linhas com dados na planilha: {stats['linhas_planilha']}")
    print(f"Sem telefone válido (ignoradas): {stats['sem_telefone']}")
    print(f"Duplicatas mescladas por telefone: {stats['duplicatas_mescladas']}")
    print(f"Contatos prontos para importar: {stats['exportados']}")
    print(f"CSV:  {csv_path}")
    print(f"XLSX: {xlsx_path}")
    if rows[:3]:
        print("Amostra:")
        for i, r in enumerate(rows[:3], 1):
            print(f"  {i}. {r['Nome']} | {r['Telefone']} | {r['Cidade']} | {r['Tags (separadas por ;)']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
